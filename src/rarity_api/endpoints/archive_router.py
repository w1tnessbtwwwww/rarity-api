from typing import Union
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
import openpyxl
from pydantic import BaseModel

from rarity_api.common.auth.native_auth.dependencies import authenticate
from rarity_api.common.auth.schemas.user import UserRead
from rarity_api.core.database.connector import get_session
from rarity_api.core.database.repos.repos import CityRepository, CountryRepository, ItemRepository, ManufacturerRepository, RegionRepository

import zipfile
import tempfile

import os
import shutil
from fastapi import HTTPException, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

router = APIRouter(
    prefix="/archive",
    tags=["archive"]
)

IMAGES_DIR = os.path.abspath("src/rarity_api/images")
print(IMAGES_DIR)  # Абсолютный путь для изображений
os.makedirs(IMAGES_DIR, exist_ok=True)

class ReadItem(BaseModel):
    rp: int | None
    country: str | None
    manufacturer_name: str | None
    region: str | None
    city: str | None
    prod_year_start: int | None
    prod_year_end: Union[int, str] | None
    desc: str | None

async def process_excel_file(session: AsyncSession, excel_path: str, source: str):
    """Обработка Excel файла и сохранение в БД"""
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    async with session.begin():  # Используем транзакцию
        for row in sheet.iter_rows(min_row=10, values_only=True):
            # Пропускаем полностью пустые строки
            if all(cell is None for cell in row):
                continue
                
            try:
                current_row = ReadItem(
                    rp=row[0],
                    country=row[1],
                    manufacturer_name=row[2],
                    region=row[3],
                    city=row[4],
                    prod_year_start=row[5],
                    prod_year_end=row[6],
                    desc=row[7]
                )
            except (TypeError, IndexError):
                continue

            # Пропускаем строки с отсутствующими обязательными данными
            if not all([current_row.rp, current_row.manufacturer_name]):
                continue

            # Репозитории
            country_repo = CountryRepository(session)
            city_repo = CityRepository(session)
            region_repo = RegionRepository(session)
            manufacturer_repo = ManufacturerRepository(session)
            item_repo = ItemRepository(session)

            # Обработка годов производства
            year_start = current_row.prod_year_start or 0
            year_end = current_row.prod_year_end or "now"
            
            # Создаем или получаем объекты
            country = await country_repo.get_or_create(name=current_row.country)
            region = await region_repo.get_or_create(
                name=current_row.region, 
                country_id=country.id
            ) if current_row.region else None
            
            city = await city_repo.get_or_create(
                name=current_row.city, 
                region_id=region.id if region else None
            ) if current_row.city else None
            
            manufacturer = await manufacturer_repo.get_or_create(
                name=current_row.manufacturer_name
            )
            
            # Создаем запись
            await item_repo.create(
                rp=current_row.rp,
                manufacturer_id=manufacturer.id,
                description=current_row.desc or "",
                production_years=f"{year_start} - {year_end}",
                city_id=city.id if city else None,
                source=source
            )

@router.post("/upload")
async def upload_archive(
    source: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            # Распаковка архива
            with zipfile.ZipFile(file.file, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
        except zipfile.BadZipFile:
            raise HTTPException(400, detail="Невозможно распаковать архив")

        # 1. Поиск Excel файла (включая index.xlsx)
        excel_path = None
        for root, _, files in os.walk(temp_dir):
            for file_name in files:
                if file_name.lower() == "index.xlsx" or file_name.endswith('.xlsx'):
                    excel_path = os.path.join(root, file_name)
                    break
            if excel_path:
                break

        if not excel_path:
            raise HTTPException(400, detail="Excel файл не найден в архиве")

        # 2. Поиск изображений (более гибкий поиск)
        images_source = None
        possible_paths = [
            os.path.join(temp_dir, "book", "marks_images"),
            os.path.join(temp_dir, "marks_images"),
            os.path.join(temp_dir, "images")
        ]
        
        for path in possible_paths:
            if os.path.exists(path) and os.path.isdir(path):
                images_source = path
                break

        if not images_source:
            raise HTTPException(400, detail="Папка с изображениями не найдена")

        # 3. Копирование изображений с очисткой целевой директории
        if os.path.exists(IMAGES_DIR):
            shutil.rmtree(IMAGES_DIR)
        shutil.copytree(images_source, IMAGES_DIR)

        # 4. Обработка Excel с явным коммитом
        try:
            await process_excel_file(session, excel_path, source)
            await session.commit()  # Явный коммит изменений
        except Exception as e:
            await session.rollback()
            raise HTTPException(500, detail=f"Ошибка обработки Excel: {str(e)}")

    return {
        "status": "success",
        "images_dir": IMAGES_DIR,
        "copied_images": len(os.listdir(IMAGES_DIR))
    }